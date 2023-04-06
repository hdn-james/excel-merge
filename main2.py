
from os import path as PATH, mkdir, listdir
from shutil import rmtree
from typing import List
from openpyxl import Workbook, load_workbook


def yes_no(message: str) -> bool:
    inp = input(message).strip()
    if inp == '':
        return True
    if inp.lower() == 'y' or inp.lower() == 'yes':
        return True
    if inp.lower() == 'n' or inp.lower() == 'no':
        return False
    return False


class Excel:
    def __init__(self) -> None:
        self._input_file_path = './input'
        self._output_file_path = './output'

        self._excel_file_list = None
        self._output_file_name = None

        self._dest_wb = Workbook()

    @property
    def input_file_path(self) -> str:
        if not self._input_file_path or self._input_file_path == '':
            raise ValueError("✕ You need to set a valid input path!")
        return self._input_file_path

    @property
    def output_file_path(self) -> str:
        if not self._output_file_path or self._output_file_path == '':
            raise ValueError("✕ You need to set a valid output path!")
        return self._output_file_path

    @property
    def excel_file_list(self) -> List[str]:
        if not self._excel_file_list or len(self._excel_file_list) == 0:
            raise ValueError("✕ You need to set a non-empty list file!")
        return self._excel_file_list

    @property
    def output_file_name(self) -> List[str]:
        if not self._output_file_name or len(self._output_file_name) == 0:
            raise ValueError("✕ You need to set a valid file name!")
        return self._output_file_name

    @input_file_path.setter
    def input_file_path(self, path: str) -> None:
        if not path or path == '':
            raise ValueError("✕ You need to set a valid input path!")
        if not PATH.exists(path):
            raise ValueError("✕ Directory does not exist!")
        self._input_file_path = path

    @output_file_path.setter
    def output_file_path(self, path: str) -> None:
        if not PATH.exists(path):
            mkdir(path=path)
        else:
            rmtree(path)
            mkdir(path)

        self._output_file_path = path

    @excel_file_list.setter
    def excel_file_list(self, list_file: List[str]) -> None:
        if not list_file or len(list_file) == 0:
            raise ValueError("✕ You need to set a non-empty list file!")
        self._excel_file_list = list_file

    @output_file_name.setter
    def output_file_name(self, file_name: str) -> None:
        if not file_name or file_name == '':
            raise ValueError("✕ You need to set a non-empty list file!")
        self._output_file_name = file_name

    def prepare_data(self) -> None:
        list_file = listdir(self._input_file_path)

        list_file = list(filter(lambda p: p.endswith('.xlsx'), list_file))

        if len(list_file) == 0:
            raise ValueError("✕ No xlsx file in folder...")

        self.excel_file_list = list_file

    def process(self) -> None:
        for excel_files in self.excel_file_list:
            sheet_name = excel_files.split('.xlsx')[0]
            self._dest_wb.create_sheet(sheet_name)
            dest_ws = self._dest_wb[sheet_name]

            source_wb = load_workbook(
                self._input_file_path + "/" + excel_files)
            source_sheet = source_wb.active
            for row in source_sheet.rows:
                for cell in row:
                    dest_ws[cell.coordinate] = cell.value

        self._dest_wb.save(self._output_file_path+self.output_file_name)

    def reset(self) -> None:
        self._excel_file_list = None
        self._input_file_path = None
        self._output_file_path = None
        self._output_file_name = None
        self._dest_wb = None
        self._dest_wb = Workbook()


if __name__ == '__main__':
    excel = Excel()

    excel.input_file_path = './input'
    excel.output_file_path = './output'
    excel.output_file_name = '/output-merged.xlsx'
    excel.prepare_data()
    excel.process()
