from os import path as PATH, mkdir, listdir
from shutil import rmtree
from typing import List
from openpyxl import Workbook, load_workbook
from copy import copy


def yes_no(message: str) -> bool:
    inp = input(message).strip()
    if inp == "":
        return True
    if inp.lower() == "y" or inp.lower() == "yes":
        return True
    if inp.lower() == "n" or inp.lower() == "no":
        return False
    return False


class Excel:
    def __init__(self) -> None:
        self._input_file_path = None
        self._output_file_path = None

        self._excel_file_list = None
        self._output_file_name = None

        self._dest_wb = Workbook()

    @property
    def input_file_path(self) -> str:
        if not self._input_file_path or self._input_file_path == "":
            raise ValueError("✕ You need to set a valid input path!")
        return self._input_file_path

    @property
    def output_file_path(self) -> str:
        if not self._output_file_path or self._output_file_path == "":
            raise ValueError("✕ You need to set a valid output path!")
        return self._output_file_path

    @property
    def excel_file_list(self) -> List[str]:
        if not self._excel_file_list or len(self._excel_file_list) == 0:
            raise ValueError("✕ You need to set a non-empty list file!")
        return self._excel_file_list

    @property
    def output_file_name(self) -> List[str]:
        if not self._output_file_name or len(self._output_file_name) == 0:
            raise ValueError("✕ You need to set a valid file name!")
        return self._output_file_name

    @input_file_path.setter
    def input_file_path(self, path: str) -> None:
        if not path or path == "":
            raise ValueError("✕ You need to set a valid input path!")
        if not PATH.exists(path):
            raise ValueError("✕ Directory does not exist!")
        self._input_file_path = path

    @output_file_path.setter
    def output_file_path(self, path: str) -> None:
        if not path or path == "":
            raise ValueError("✕ You need to set a valid output path!")
        if not PATH.exists(path):
            opt = yes_no("⚠ Output folder does not exist! Create a new one? (Y/n): ")
            if opt:
                mkdir(path=path)
            else:
                exit(0)
        else:
            options = yes_no("⚠ Old data will be deleted. Continue? (Y/n): ")
            if options:
                rmtree(path)
                mkdir(path)
            else:
                exit(0)
        self._output_file_path = path

    @excel_file_list.setter
    def excel_file_list(self, list_file: List[str]) -> None:
        if not list_file or len(list_file) == 0:
            raise ValueError("✕ You need to set a non-empty list file!")
        self._excel_file_list = list_file

    @output_file_name.setter
    def output_file_name(self, file_name: str) -> None:
        if not file_name or file_name == "":
            raise ValueError("✕ You need to set a non-empty list file!")
        self._output_file_name = file_name

    def prepare_data(self) -> None:
        list_file = listdir(self._input_file_path)

        list_file = list(filter(lambda p: p.endswith(".xlsx"), list_file))

        if len(list_file) == 0:
            raise ValueError("✕ No xlsx file in folder...")

        self.excel_file_list = list_file

    def copy_sheet(self, source, dest):
        pass

    def copy_sheet_attributes(self, source, dest):
        pass

    def process(self) -> None:
        for excel_files in self.excel_file_list:
            sheet_name = excel_files.split(".xlsx")[0]
            self._dest_wb.create_sheet(sheet_name)
            dest_ws = self._dest_wb[sheet_name]

            source_wb = load_workbook(self._input_file_path + "/" + excel_files)
            source_sheet = source_wb.active
            # dest_ws = source_wb.copy_worksheet(source_sheet)
            for row in source_sheet.rows:
                for cell in row:
                    dest_ws[cell.coordinate] = cell.value
                    if cell.has_style:
                        dest_ws[cell.coordinate].font = copy(cell.font)
                        dest_ws[cell.coordinate].border = copy(cell.border)
                        dest_ws[cell.coordinate].fill = copy(cell.fill)
                        dest_ws[cell.coordinate].number_format = copy(
                            cell.number_format
                        )
                        dest_ws[cell.coordinate].protection = copy(cell.protection)
                        dest_ws[cell.coordinate].alignment = copy(cell.alignment)

                    if cell.hyperlink:
                        dest_ws[cell.coordinate]._hyperlink = copy(cell.hyperlink)

                    if cell.comment:
                        dest_ws[cell.coordinate].comment = copy(cell.comment)

            dest_ws.sheet_format = copy(source_sheet.sheet_format)
            dest_ws.sheet_properties = copy(source_sheet.sheet_properties)
            dest_ws.merged_cells = copy(source_sheet.merged_cells)
            dest_ws.page_margins = copy(source_sheet.page_margins)
            dest_ws.freeze_panes = copy(source_sheet.freeze_panes)

            # set row dimensions
            # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
            for rn in range(len(source_sheet.row_dimensions)):
                dest_ws.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

            if source_sheet.sheet_format.defaultColWidth is None:
                print("Unable to copy default column wide:::", sheet_name)
            else:
                dest_ws.sheet_format.defaultColWidth = copy(
                    source_sheet.sheet_format.defaultColWidth
                )

            # set specific column width and hidden property
            # we cannot copy the entire column_dimensions attribute so we copy selected attributes
            for key, value in source_sheet.column_dimensions.items():
                dest_ws.column_dimensions[key].min = copy(
                    source_sheet.column_dimensions[key].min
                )  # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
                dest_ws.column_dimensions[key].max = copy(
                    source_sheet.column_dimensions[key].max
                )  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
                dest_ws.column_dimensions[key].width = copy(
                    source_sheet.column_dimensions[key].width
                )  # set width for every column
                dest_ws.column_dimensions[key].hidden = copy(
                    source_sheet.column_dimensions[key].hidden
                )

        if 'Sheet' in self._dest_wb.sheetnames:  # remove default sheet
            self._dest_wb.remove(self._dest_wb['Sheet'])
        self._dest_wb.save(self._output_file_path + self.output_file_name)

    def reset(self) -> None:
        self._excel_file_list = None
        self._input_file_path = None
        self._output_file_path = None
        self._output_file_name = None
        self._dest_wb = None
        self._dest_wb = Workbook()
