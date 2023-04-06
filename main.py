from excel import Excel, yes_no
from os import system, name, get_terminal_size
from colorama import Fore, Style


from os import path as PATH, mkdir, listdir
from shutil import rmtree
from typing import List
from openpyxl import Workbook, load_workbook


def yes_no(message: str) -> bool:
    inp = input(message).strip()
    if inp == '':
        return True
    if inp.lower() == 'y' or inp.lower() == 'yes':
        return True
    if inp.lower() == 'n' or inp.lower() == 'no':
        return False
    return False


class Excel:
    def __init__(self) -> None:
        self._input_file_path = None
        self._output_file_path = None

        self._excel_file_list = None
        self._output_file_name = None

        self._dest_wb = Workbook()

    @property
    def input_file_path(self) -> str:
        if not self._input_file_path or self._input_file_path == '':
            raise ValueError("✕ You need to set a valid input path!")
        return self._input_file_path

    @property
    def output_file_path(self) -> str:
        if not self._output_file_path or self._output_file_path == '':
            raise ValueError("✕ You need to set a valid output path!")
        return self._output_file_path

    @property
    def excel_file_list(self) -> List[str]:
        if not self._excel_file_list or len(self._excel_file_list) == 0:
            raise ValueError("✕ You need to set a non-empty list file!")
        return self._excel_file_list

    @property
    def output_file_name(self) -> List[str]:
        if not self._output_file_name or len(self._output_file_name) == 0:
            raise ValueError("✕ You need to set a valid file name!")
        return self._output_file_name

    @input_file_path.setter
    def input_file_path(self, path: str) -> None:
        if not path or path == '':
            raise ValueError("✕ You need to set a valid input path!")
        if not PATH.exists(path):
            raise ValueError("✕ Directory does not exist!")
        self._input_file_path = path

    @output_file_path.setter
    def output_file_path(self, path: str) -> None:
        if not path or path == '':
            raise ValueError("✕ You need to set a valid output path!")
        if not PATH.exists(path):
            opt = yes_no(
                '⚠ Output folder does not exist! Create a new one? (Y/n): ')
            if opt:
                mkdir(path=path)
            else:
                exit(0)
        else:
            options = yes_no("⚠ Old data will be deleted. Continue? (Y/n): ")
            if options:
                rmtree(path)
                mkdir(path)
            else:
                exit(0)
        self._output_file_path = path

    @excel_file_list.setter
    def excel_file_list(self, list_file: List[str]) -> None:
        if not list_file or len(list_file) == 0:
            raise ValueError("✕ You need to set a non-empty list file!")
        self._excel_file_list = list_file

    @output_file_name.setter
    def output_file_name(self, file_name: str) -> None:
        if not file_name or file_name == '':
            raise ValueError("✕ You need to set a non-empty list file!")
        self._output_file_name = file_name

    def prepare_data(self) -> None:
        list_file = listdir(self._input_file_path)

        list_file = list(filter(lambda p: p.endswith('.xlsx'), list_file))

        if len(list_file) == 0:
            raise ValueError("✕ No xlsx file in folder...")

        self.excel_file_list = list_file

    def process(self) -> None:
        for excel_files in self.excel_file_list:
            sheet_name = excel_files.split('.xlsx')[0]
            self._dest_wb.create_sheet(sheet_name)
            dest_ws = self._dest_wb[sheet_name]

            source_wb = load_workbook(
                self._input_file_path + "/" + excel_files)
            source_sheet = source_wb.active
            for row in source_sheet.rows:
                for cell in row:
                    dest_ws[cell.coordinate] = cell.value

        self._dest_wb.save(self._output_file_path+self.output_file_name)

    def reset(self) -> None:
        self._excel_file_list = None
        self._input_file_path = None
        self._output_file_path = None
        self._output_file_name = None
        self._dest_wb = None
        self._dest_wb = Workbook()


def clear_screen() -> None:
    if name == 'nt':
        system('cls')
    else:
        system('clear')


class UI:
    def __init__(self) -> None:  # 73 cols
        self._ex = Excel()
        self.cols, self.row = get_terminal_size()
        self.running = True
        self._step = 1

    def terminal_size(self) -> None:
        self.cols, self.row = get_terminal_size()

    def welcome(self) -> None:
        to_print = '''
                      /^--^\     /^--^\     /^--^\\
                      \____/     \____/     \____/
                     /      \   /      \   /      \\
                    |        | |        | |        |
                     \__  __/   \__  __/   \__  __/      huynhdainhan
|^|^|^|^|^|^|^|^|^|^|^|^\ \^|^|^|^/ /^|^|^|^|^\ \^|^|^|^|^|^|^|^|^|^|^|^|
| | | | | | | | | | | | |\ \| | |/ /| | | | | | \ \ | | | | | | | | | | |
| | | | | | | | | | | | / / | | |\ \| | | | | |/ /| | | | | | | | | | | |
| | | | | | | | | | | | \/| | | | \/| | | | | |\/ | | | | | | | | | | | |
#########################################################################
| | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | |
| | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | | |

'''
        print(Fore.YELLOW, to_print, Style.RESET_ALL)

    def options(self, step) -> None:
        if step == 1:
            while True:
                try:
                    inp = input("Step 1: Input data folder: ").strip()
                    if not inp.endswith('/'):
                        inp = inp + '/'
                    self._ex.input_file_path = inp
                    self._step = 2
                    break
                except ValueError as error:
                    print(Fore.RED, error, Style.RESET_ALL)

        if step == 2:
            while True:
                try:
                    print(Fore.GREEN, "Input folder:",
                          self._ex.input_file_path, Style.RESET_ALL)
                    inp = input("Step 2: Output data folder: ").strip()
                    if not inp.endswith('/'):
                        inp = inp + '/'
                    self._ex.output_file_path = inp
                    self._step = 3
                    break
                except ValueError as error:
                    print(Fore.RED, error, Style.RESET_ALL)

        if step == 3:
            while True:
                try:
                    print(Fore.GREEN, "Input folder:",
                          self._ex.input_file_path, Style.RESET_ALL)
                    print(Fore.GREEN, "Output folder:",
                          self._ex.output_file_path, Style.RESET_ALL)
                    inp = input(
                        "Step 3: Name your output file (.xlsx): ").strip()
                    if not inp.endswith('.xlsx'):
                        inp = inp + '.xlsx'
                    self._ex.output_file_name = inp
                    self._step = 4
                    break
                except ValueError as error:
                    print(Fore.RED, error, Style.RESET_ALL)

        if step == 4:
            try:
                print(Fore.LIGHTGREEN_EX,
                      "ⓘ Loading data...", Style.RESET_ALL)
                self._ex.prepare_data()
                print(Fore.LIGHTGREEN_EX,
                      "ⓘ List file will be merged", Style.RESET_ALL)
                for idx, file in enumerate(self._ex.excel_file_list):
                    print(Fore.YELLOW, idx+1, '-\t', file, Style.RESET_ALL)

                opt = yes_no("Continue? (Y/n): ")

                if opt:
                    self._step = 5
                else:
                    self.running = False
            except ValueError as error:
                print(error)
                opt = yes_no("Try again? (Y/n): ")

                if opt:
                    self._step = 1
                else:
                    self.running = False
            except:
                opt = yes_no("Something went wrong. Try again? (Y/n): ")

                if opt:
                    self._step = 1
                else:
                    self.running = False

        if step == 5:
            try:
                print(Fore.LIGHTGREEN_EX,
                      "ⓘ Processing data...", Style.RESET_ALL)
                self._ex.process()
                print(Fore.LIGHTGREEN_EX,
                      "ⓘ Result file saved to:", self._ex.output_file_path+self._ex.output_file_name, Style.RESET_ALL)

                opt = yes_no("Wanna merge more files? (Y/n): ")

                if opt:
                    self._step = 1
                    self._ex.reset()
                else:
                    self.running = False
            except:
                opt = yes_no("Something went wrong. Try again? (Y/n): ")

                if opt:
                    self._step = 1
                else:
                    self.running = False

    def run(self) -> None:
        while (self.running):
            clear_screen()

            if not self.cols > 73 and not self.row > 27:
                print(Fore.BLUE, "Columns:", self.cols,
                      "x Rows:", self.row, Style.RESET_ALL)
                print(
                    Fore.YELLOW, "⚠ Your window size is too small...\n Columns must > 73 and Rows must > 27", Style.RESET_ALL)
                self.terminal_size()
                continue

            self.welcome()
            self.options(self._step)


if __name__ == '__main__':
    u = UI()

    u.run()
